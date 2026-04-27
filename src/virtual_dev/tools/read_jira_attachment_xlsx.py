"""Download an XLSX attached to a Jira ticket and return its rows.

The actual ticket DM-3342 has an XLSX attachment (Манжерок trasses),
so this is the same pattern as PDF/DOCX but for spreadsheets — flatten
each sheet to tab-separated rows so the LLM can read it inline.
"""

from __future__ import annotations

import asyncio
import io
from typing import Any

from claude_agent_sdk import tool
from loguru import logger
from openpyxl import load_workbook

from virtual_dev.tools import ToolContext
from virtual_dev.tools._helpers import (
    error_text,
    fetch_jira_attachment_content,
    parse_jira_attachment_id,
    text_result,
)
from virtual_dev.tools.read_jira_attachment_pdf import _wrap_untrusted

TOOL_GROUP = "researcher"

_DEFAULT_MAX_CHARS = 30_000


def build(ctx: ToolContext):
    if ctx.settings is None:
        return None
    settings = ctx.settings

    @tool(
        "read_jira_attachment_xlsx",
        "Download an XLSX attached to a Jira ticket and return its "
        "rows as tab-separated text, one block per sheet. Pass the "
        "full URL as it appears in the ticket. Authenticated with "
        "the Jira PAT. Output is wrapped as untrusted content. "
        "Truncates at max_chars (default 30000) — for large sheets "
        "consider asking a human for a CSV export instead.",
        {
            "type": "object",
            "properties": {
                "url": {"type": "string"},
                "max_chars": {"type": "integer"},
            },
            "required": ["url"],
        },
    )
    async def _read_xlsx(args: dict[str, Any]) -> dict[str, Any]:
        return await run(settings, args)

    return _read_xlsx


async def run(settings, args: dict[str, Any]) -> dict[str, Any]:
    url = str(args.get("url") or "").strip()
    max_chars = int(args.get("max_chars") or _DEFAULT_MAX_CHARS)
    if not url:
        return error_text("Empty URL")
    if not settings.jira_url or not settings.jira_token:
        return error_text("Jira credentials are not configured (JIRA_URL / JIRA_TOKEN)")
    attachment_id = parse_jira_attachment_id(url)
    if attachment_id is None:
        return error_text(
            f"Couldn't extract an attachment id from {url!r}. Expected "
            f"either a /secure/attachment/<id>/<filename> URL or a bare "
            f"numeric id."
        )
    try:
        body = await asyncio.to_thread(
            fetch_jira_attachment_content,
            jira_url=settings.jira_url,
            jira_token=settings.jira_token,
            attachment_id=attachment_id,
        )
    except Exception as exc:
        logger.exception("read_jira_attachment_xlsx: download failed")
        return error_text(f"Download failed (attachment {attachment_id}): {exc}")

    try:
        parts = await asyncio.to_thread(_xlsx_to_text, body)
    except Exception as exc:
        logger.exception("read_jira_attachment_xlsx: XLSX parse failed")
        return error_text(f"XLSX parse failed: {exc}")

    full = "\n\n".join(parts)
    truncated = False
    if len(full) > max_chars:
        full = full[:max_chars] + f"\n... [truncated, {len(full)} chars total]"
        truncated = True
    header = f"# XLSX: {url}" + (" (truncated)" if truncated else "") + "\n\n"
    return text_result(_wrap_untrusted(header + full, source=f"jira:xlsx:{url}"))


def _xlsx_to_text(body: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    blocks: list[str] = []
    for sheet in wb.worksheets:
        rows: list[str] = [f"## sheet: {sheet.title}"]
        for raw_row in sheet.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in raw_row]
            # Skip rows that are entirely blank — common in xlsx.
            if any(c.strip() for c in cells):
                rows.append("\t".join(cells))
        blocks.append("\n".join(rows))
    return blocks
