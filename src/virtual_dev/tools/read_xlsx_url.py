"""Download an XLSX from any URL and return its rows.

Same host-aware auth dispatch as ``read_pdf_url`` — works for
Jira / Mattermost / Confluence attachments and public URLs.
"""

from __future__ import annotations

import asyncio
import io
from typing import Any

from claude_agent_sdk import tool
from loguru import logger
from openpyxl import load_workbook

from virtual_dev.tools import ToolContext
from virtual_dev.tools._helpers import download_url_bytes, error_text, text_result
from virtual_dev.tools.read_pdf_url import _wrap_untrusted

TOOL_GROUP = "shared"

_DEFAULT_MAX_CHARS = 30_000


def build(ctx: ToolContext):
    if ctx.settings is None:
        return None
    settings = ctx.settings

    @tool(
        "read_xlsx_url",
        "Download an XLSX from any URL and return its rows as "
        "tab-separated text, one block per sheet. Works for Jira / "
        "Mattermost / Confluence attachments and public URLs — auth "
        "is host-aware. Pass the full URL exactly. Output is wrapped "
        "as untrusted content. Truncates at max_chars (default 30000) "
        "— for large sheets consider asking a human for a CSV export "
        "instead.",
        {
            "type": "object",
            "properties": {
                "url": {"type": "string"},
                "max_chars": {"type": "integer"},
            },
            "required": ["url"],
        },
    )
    async def _read_xlsx(args: dict[str, Any]) -> dict[str, Any]:
        return await run(settings, args)

    return _read_xlsx


async def run(settings, args: dict[str, Any]) -> dict[str, Any]:
    url = str(args.get("url") or "").strip()
    max_chars = int(args.get("max_chars") or _DEFAULT_MAX_CHARS)
    if not url:
        return error_text("Empty URL")

    try:
        body = await asyncio.to_thread(download_url_bytes, url, settings)
    except Exception as exc:
        logger.exception("read_xlsx_url: download failed for {}", url)
        return error_text(f"Download failed: {exc}")

    try:
        parts = await asyncio.to_thread(_xlsx_to_text, body)
    except Exception as exc:
        logger.exception("read_xlsx_url: XLSX parse failed for {}", url)
        return error_text(f"XLSX parse failed: {exc}")

    full = "\n\n".join(parts)
    truncated = False
    if len(full) > max_chars:
        full = full[:max_chars] + f"\n... [truncated, {len(full)} chars total]"
        truncated = True
    header = f"# XLSX: {url}" + (" (truncated)" if truncated else "") + "\n\n"
    return text_result(_wrap_untrusted(header + full, source=f"xlsx:{url}"))


def _xlsx_to_text(body: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    blocks: list[str] = []
    for sheet in wb.worksheets:
        rows: list[str] = [f"## sheet: {sheet.title}"]
        for raw_row in sheet.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in raw_row]
            # Skip rows that are entirely blank — common in xlsx.
            if any(c.strip() for c in cells):
                rows.append("\t".join(cells))
        blocks.append("\n".join(rows))
    return blocks
